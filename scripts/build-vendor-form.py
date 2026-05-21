#!/usr/bin/env python3
"""生成「ETF 厂商产品提报表」Excel 模板"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

FONT = '微软雅黑'
GOLD = 'D4A017'
DARK = '1A1A1A'
LIGHT = 'FFF6E0'
GREY = '888888'

thin = Side(style='thin', color='CCCCCC')
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ============ Sheet 1: 填写说明 ============
ws1 = wb.active
ws1.title = '填写说明'
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 4
ws1.column_dimensions['B'].width = 96

def line(row, text, *, bold=False, size=11, color='000000', bg=None, indent=0):
    c = ws1.cell(row=row, column=2, value=text)
    c.font = Font(name=FONT, bold=bold, size=size, color=color)
    c.alignment = Alignment(wrap_text=True, vertical='center', indent=indent)
    if bg:
        c.fill = PatternFill('solid', start_color=bg)
    return c

ws1.cell(row=1, column=2, value='ETF 厂商产品提报表 · 填写说明')
ws1['B1'].font = Font(name=FONT, bold=True, size=18, color=GOLD)
ws1.row_dimensions[1].height = 36

rows1 = [
    ('「做 ETF 球队老板」H5 活动 — 招商产品提报', dict(bold=True, size=12, color=DARK)),
    ('', {}),
    ('一、活动背景', dict(bold=True, size=12, color=GOLD)),
    ('用户在 H5 中通过选阵型挑选 11 只 ETF 组成"球队"。每个位置（前锋/中场/后卫/门将）有一个产品池，', {}),
    ('贵司投放的 ETF 将进入对应位置池，并通过"跑马灯轮转"机制获得头部曝光（每日轮到第一位享"今日推荐"金徽章）。', {}),
    ('', {}),
    ('二、提报方式', dict(bold=True, size=12, color=GOLD)),
    ('请在「产品提报表」工作表中填写贵司要投放的 ETF 产品，每行一只。', {}),
    ('每周可更换一批产品，运营同学会按提报表手动更新到 H5。', {}),
    ('', {}),
    ('三、字段说明', dict(bold=True, size=12, color=GOLD)),
    ('• 位置：必填。从下拉选择 前锋 / 中场 / 后卫 / 门将。建议按产品风格归位：', {}),
    ('     前锋 = 高弹性进攻型（科技/成长/主题）； 中场 = 宽基/核心配置；', dict(color=GREY)),
    ('     后卫 = 防守稳健（红利/低波/价值）； 门将 = 避险压舱（黄金/债券/货币）。', dict(color=GREY)),
    ('• ETF简称：必填。展示在卡片和海报上的名称，建议 ≤ 10 字。', {}),
    ('• ETF代码：必填。6 位交易代码（如 512710）。', {}),
    ('• 基金公司：必填。用于跑马灯轮转分组（同公司产品归为一家）。', {}),
    ('• 产品一句话简介：必填，≤ 30 字。展示在卡片第二行，介绍产品特点。', {}),
    ('• 风险等级：必填。从下拉选择 极低 / 低 / 低中 / 中低 / 中 / 中高 / 高。', {}),
    ('• 建议球员技能名：选填。趣味标签（如"爆发突击""稳健守护"），不填则由运营拟定。', {}),
    ('• 备注：选填。其他需要说明的信息。', {}),
    ('', {}),
    ('四、注意事项', dict(bold=True, size=12, color=GOLD)),
    ('• 五维评分（进攻力/防守力/控场力/平衡度/个性值）由雪球运营根据产品类型统一设定，厂商无需填写。', {}),
    ('• 近一周涨跌幅由系统对接行情接口自动获取，厂商无需填写。', {}),
    ('• 同一基金公司可在多个位置投放产品；建议每个位置至少提报 1 只，覆盖更全。', {}),
    ('• 标黄行为示例，请勿修改；从空白行开始填写贵司产品。', dict(color='C00000')),
]
r = 3
for text, style in rows1:
    line(r, text, **style)
    ws1.row_dimensions[r].height = 22 if text else 10
    r += 1

# ============ Sheet 2: 产品提报表 ============
ws2 = wb.create_sheet('产品提报表')
ws2.sheet_view.showGridLines = False

headers = ['序号', '位置', 'ETF简称', 'ETF代码', '基金公司',
           '产品一句话简介（≤30字）', '风险等级', '建议球员技能名', '备注']
widths = [6, 10, 16, 12, 16, 38, 12, 16, 24]
for i, w in enumerate(widths, start=1):
    ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

# 标题行
ws2.merge_cells('A1:I1')
t = ws2['A1']
t.value = 'ETF 厂商产品提报表（「做 ETF 球队老板」H5 活动）'
t.font = Font(name=FONT, bold=True, size=14, color=GOLD)
t.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 32

ws2.merge_cells('A2:I2')
sub = ws2['A2']
sub.value = '填写前请先阅读「填写说明」工作表 ｜ 标黄行为示例 ｜ 带 * 为必填'
sub.font = Font(name=FONT, size=10, color=GREY)
sub.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[2].height = 20

# 表头
HEAD_ROW = 3
required = {'位置', 'ETF简称', 'ETF代码', '基金公司', '产品一句话简介（≤30字）', '风险等级'}
for i, h in enumerate(headers, start=1):
    c = ws2.cell(row=HEAD_ROW, column=i, value=h + (' *' if h in required else ''))
    c.font = Font(name=FONT, bold=True, size=11, color='FFFFFF')
    c.fill = PatternFill('solid', start_color=DARK)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = box
ws2.row_dimensions[HEAD_ROW].height = 30

# 示例行
examples = [
    ('前锋', '军工龙头ETF', '512710', '富国基金', '聚焦军工龙头，国防主题进攻锋线，事件驱动弹性强。', '高', '尖刀突袭', '示例，请勿修改'),
    ('中场', '消费50ETF', '515650', '富国基金', '聚焦消费 50 龙头，攻守兼备的中场组织者。', '中', '消费节拍', '示例，请勿修改'),
    ('后卫', '上证国债ETF', '511220', '富国基金', '配置上证 5 年期以上国债，安全垫极厚。', '极低', '钢铁后防', '示例，请勿修改'),
    ('门将', '短融ETF', '511360', '富国基金', '聚焦短期债券，收益稳健、流动性好。', '极低', '关键封堵', '示例，请勿修改'),
]
row = HEAD_ROW + 1
for idx, ex in enumerate(examples, start=1):
    ws2.cell(row=row, column=1, value=idx)
    for col, val in enumerate(ex, start=2):
        ws2.cell(row=row, column=col, value=val)
    for col in range(1, 10):
        cell = ws2.cell(row=row, column=col)
        cell.font = Font(name=FONT, size=10, color=GREY, italic=True)
        cell.fill = PatternFill('solid', start_color='FFF6E0')
        cell.border = box
        cell.alignment = Alignment(vertical='center', wrap_text=True,
                                   horizontal='center' if col in (1, 2, 4, 7) else 'left')
    ws2.row_dimensions[row].height = 30
    row += 1

# 空白填写行
FIRST_BLANK = row
BLANK_ROWS = 30
for i in range(BLANK_ROWS):
    ws2.cell(row=row, column=1, value=i + 1)
    for col in range(1, 10):
        cell = ws2.cell(row=row, column=col)
        cell.font = Font(name=FONT, size=10)
        cell.border = box
        cell.alignment = Alignment(vertical='center', wrap_text=True,
                                   horizontal='center' if col in (1, 2, 4, 7) else 'left')
    ws2.row_dimensions[row].height = 28
    row += 1
LAST_ROW = row - 1

# 数据验证下拉
dv_pos = DataValidation(type='list', formula1='"前锋,中场,后卫,门将"', allow_blank=True)
dv_pos.error = '请从下拉选择：前锋 / 中场 / 后卫 / 门将'
dv_pos.prompt = '从下拉选择位置'
ws2.add_data_validation(dv_pos)
dv_pos.add(f'B{HEAD_ROW + 1}:B{LAST_ROW}')

dv_risk = DataValidation(type='list', formula1='"极低,低,低中,中低,中,中高,高"', allow_blank=True)
dv_risk.error = '请从下拉选择风险等级'
dv_risk.prompt = '从下拉选择风险等级'
ws2.add_data_validation(dv_risk)
dv_risk.add(f'G{HEAD_ROW + 1}:G{LAST_ROW}')

# 代码列文本格式（避免 512710 被当数字）
for rr in range(HEAD_ROW + 1, LAST_ROW + 1):
    ws2.cell(row=rr, column=4).number_format = '@'

ws2.freeze_panes = f'A{HEAD_ROW + 1}'

wb.save('/Users/liujiangtao/Desktop/etf世界杯/ETF厂商产品提报表.xlsx')
print('✓ written ETF厂商产品提报表.xlsx')
